import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

data = {
    "分类": ["新房", "新房", "新房", "二手房", "二手房", "二手房"],
    "梯队": ["一线", "二线", "三四线", "一线", "二线", "三四线"],
    "48周": [93, 285, 62, 13985, 16244, 2940],
    "49周": [63, 197, 54, 13929, 24376, 3966],
    "50周": [95, 194, 58, 13492, 19444, 3708],
    "51周": [82, 251, 58, 12346, 18048, 3579],
    "51周环比": ["-13.71%", "29.54%", "-0.34%", "-8.49%", "-7.18%", "-3.48%"],
    "51周同比": ["59.41%", "65.39%", "31.11%", "115.73%", "108.29%", "77.53%"],
}

# 创建DataFrame
df = pd.DataFrame(data)

# 保存为Excel文件
file_name = "custom_table.xlsx"
with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
    # 写入数据
    df.to_excel(writer, index=False, startrow=2, sheet_name="Sheet1")

    # 获取工作表
    workbook = writer.book
    worksheet = writer.sheets["Sheet1"]

    # 定制表头
    worksheet.merge_cells("A1:A2")
    worksheet.merge_cells("B1:B2")
    worksheet.merge_cells("C1:F1")
    worksheet.merge_cells("G1:G2")
    worksheet.merge_cells("H1:H2")

    worksheet["A1"] = "分类"
    worksheet["B1"] = "梯队"
    worksheet["C1"] = "成交规模（万m2、套）"
    worksheet["G1"] = "51周\n环比"
    worksheet["H1"] = "51周\n同比"

    # 调整表头格式
    for cell in ["A1", "B1", "C1", "G1", "H1"]:
        worksheet[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet[cell].font = Font(bold=True)

    for col in range(3, 7):
        worksheet.cell(row=2, column=col).value = f"{47 + col}周"
        worksheet.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")
        worksheet.cell(row=2, column=col).font = Font(bold=True)

    # 合并"新房"和"二手房"单元格
    worksheet.merge_cells("A3:A5")
    worksheet.merge_cells("A6:A8")

    worksheet["A3"].alignment = Alignment(vertical="center", horizontal="center")
    worksheet["A6"].alignment = Alignment(vertical="center", horizontal="center")

    # 修改梯队名称
    worksheet["B3"].value = "一线\n（4个）"
    worksheet["B4"].value = "二线\n（12个）"
    worksheet["B5"].value = "三四线\n（14个）"
    worksheet["B6"].value = "一线\n（3个）"
    worksheet["B7"].value = "二线\n（10个）"
    worksheet["B8"].value = "三四线\n（7个）"

    for row in range(3, 9):
        worksheet[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 设置51周环比和同比字体颜色
    for row, value in enumerate(df["51周环比"], start=3):
        cell = worksheet[f"G{row}"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.value = value
        if float(value.strip('%')) < 0:
            cell.font = Font(color="FF0000")  # 红色
        else:
            cell.font = Font(color="008000")  # 绿色

    for row, value in enumerate(df["51周同比"], start=3):
        cell = worksheet[f"H{row}"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.value = value
        if float(value.strip('%')) < 0:
            cell.font = Font(color="FF0000")  # 红色
        else:
            cell.font = Font(color="008000")  # 绿色

print(f"文件已保存为 {file_name}")
